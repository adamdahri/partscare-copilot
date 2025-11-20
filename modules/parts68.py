# modules/parts68.py
import os, re, fitz, openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment

# ==================== CONFIG ====================
ATA_MIN, ATA_MAX = 21, 79
DATE_RE = r"\b\d{1,2}[-\s]?(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[A-Za-z]*[-\s]?\d{2,4}\b"
PN_RE = re.compile(r"\b(?:P\/N|PN|PART\s*NO\.?|PART\s*NUMBER)[:=\s\-]*([A-Z0-9][A-Z0-9\-\/\.]+)", re.I)
HRS_RE = re.compile(r"\bHRS[:=\s\-]*([0-9.,:]+)\b", re.I)
MOS_RE = re.compile(r"\bMOS(?:\/MSC)?\b", re.I)
ATA_REF_RE = re.compile(r"^\s*(\d{5,6})\b")

# ==================== HELPERS ====================
def _norm(s: str) -> str:
    import re
    return re.sub(r"\s+", " ", s or "").strip()

def _parse_date_any(s: str) -> str:
    if not s:
        return ""
    s = s.replace("/", "-").replace(".", "-")
    s = re.sub(r"\s+", "-", s).strip()
    for f in ["%d-%b-%Y", "%d-%b-%y", "%d-%B-%Y", "%d-%B-%y", "%d-%m-%Y", "%d-%m-%y"]:
        try:
            return datetime.strptime(s.title(), f).strftime("%d-%b-%Y")
        except:
            pass
    return ""

def _hours_to_float(s: str) -> float:
    s = s.replace(",", ".").strip()
    m = re.match(r"^(\d+):(\d{2})$", s)
    if m:
        return float(m.group(1)) + float(m.group(2)) / 60.0
    try:
        return float(re.sub(r"[^\d.]", "", s))
    except:
        return 0.0

# ==================== FABRICATION DATE FIX ====================
def _detect_fabrication_date(text):
    """
    Zoekt expliciet naar de CERTIFICATION DATE in het AIRCRAFT-blok.
    Als niet gevonden, fallback naar oudste datum.
    """
    m_aircraft = re.search(r"AIRCRAFT.*?CERTIFICATION\s*DATE[:\s\-]*" + DATE_RE, text, re.I | re.S)
    if m_aircraft:
        d = re.search(DATE_RE, m_aircraft.group(0), re.I)
        if d:
            parsed = _parse_date_any(d.group(0))
            if parsed:
                return parsed

    m = re.search(r"CERTIFICATION\s*DATE[:\s\-]*" + DATE_RE, text, re.I)
    if m:
        d = re.search(DATE_RE, m.group(0), re.I)
        if d:
            parsed = _parse_date_any(d.group(0))
            if parsed:
                return parsed

    dates = []
    for m in re.finditer(DATE_RE, text, re.I):
        dt = _parse_date_any(m.group(0))
        if dt:
            try:
                dates.append(datetime.strptime(dt, "%d-%b-%Y"))
            except:
                pass
    return min(dates).strftime("%d-%b-%Y") if dates else ""

# ==================== CORE PARSING ====================
def _extract_components(pdf_path):
    with fitz.open(pdf_path) as doc:
        first_text = "\n".join(doc[p].get_text("text") for p in range(min(5, len(doc))))
        fab_date = _detect_fabrication_date(first_text)

        all_lines = []
        for p in doc:
            blocks = p.get_text("blocks")
            page_lines = []
            for b in sorted(blocks, key=lambda x: (x[1], x[0])):
                text = _norm(b[4])
                if text:
                    page_lines.append(text)
            all_lines.extend(page_lines)

    blocks = []
    cur = None
    for ln in all_lines:
        m = ATA_REF_RE.match(ln)
        if m:
            if cur:
                blocks.append(cur)
            cur = {"ref": m.group(1), "head": _norm(ln[m.end():]), "lines": []}
        elif cur:
            cur["lines"].append(ln)
    if cur:
        blocks.append(cur)

    results = []
    for b in blocks:
        ref = b["ref"]
        ata_code = ref[:2]
        if not (ata_code.isdigit() and ATA_MIN <= int(ata_code) <= ATA_MAX):
            continue

        buf = " ".join([b["head"]] + b["lines"])
        pn_match = PN_RE.search(buf)
        hrs_match = HRS_RE.search(buf)
        date_match = re.search(DATE_RE, buf, re.I)

        if not date_match:
            for l in b["lines"]:
                if MOS_RE.search(l):
                    date_match = re.search(DATE_RE, l, re.I)
                    if date_match:
                        break

        if not pn_match or not hrs_match or not date_match:
            continue

        pn = pn_match.group(1).strip()
        hrs = _hours_to_float(hrs_match.group(1))
        if hrs <= 0:
            continue

        date_val = _parse_date_any(date_match.group(0))
        if fab_date and date_val == fab_date:
            continue

        desc = re.sub(PN_RE, "", b["head"])
        desc = re.sub(HRS_RE, "", desc)
        desc = re.sub(DATE_RE, "", desc, flags=re.I)
        desc = _norm(desc)

        results.append((ata_code, ref, pn, desc, date_val, str(hrs)))

    return fab_date, results

# ==================== EXCEL EXPORT ====================
def _export_excel(records, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Components"
    headers = ["ATA CODE", "ATA REF", "PART NUMBER", "DESCRIPTION", "DATE", "HRS INFO"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    alt_fill = PatternFill("solid", fgColor="F2F2F2")
    for i, row in enumerate(records, start=2):
        ws.append(row)
        if i % 2 == 0:
            for c in ws[i]:
                c.fill = alt_fill

    for col in ws.columns:
        width = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(width + 3, 50)

    wb.save(out_path)

# ==================== PUBLIC API (voor de Copilot GUI) ====================
def run_batch(input_dir: str, output_dir: str, log=print) -> dict:
    """
    Verwerkt alle PDF's in input_dir en schrijft per PDF een Excel naar output_dir.
    log: callable(str) voor GUI-logging.
    Returnt dict met stats: {'processed': N, 'exported': M}
    """
    if not os.path.isdir(input_dir):
        raise ValueError("Inputmap bestaat niet.")
    if not os.path.isdir(output_dir):
        raise ValueError("Outputmap bestaat niet.")

    pdfs = [f for f in os.listdir(input_dir) if f.lower().endswith(".pdf")]
    if not pdfs:
        raise ValueError("Geen PDF-bestanden in inputmap.")

    exported = 0
    for i, name in enumerate(sorted(pdfs), start=1):
        log(f"\n🔍 [{i}/{len(pdfs)}] {name}")
        try:
            fab, recs = _extract_components(os.path.join(input_dir, name))
            log(f"   • Certification Date: {fab or 'niet gevonden'}")
            log(f"   • Records (na filter): {len(recs)}")
            if recs:
                out_path = os.path.join(output_dir, os.path.splitext(name)[0].replace(" ", "_") + ".xlsx")
                _export_excel(recs, out_path)
                log(f"   ✅ Export voltooid → {os.path.basename(out_path)}")
                exported += 1
            else:
                log("   ⚠️ Geen onderdelen gevonden.")
        except Exception as e:
            log(f"   ❌ Fout: {e}")

    log("\n🎯 Alles verwerkt.")
    return {"processed": len(pdfs), "exported": exported}
